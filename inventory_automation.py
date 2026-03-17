import os
import shutil
import glob
import traceback
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# One drive folder paths that get replaced daily with the morning email.
reports_folder = r"C:\Users\Administrator\OneDrive - Voxx Products\Desktop\Reports"
# Current master data path
master_data_path = r"C:\Users\Administrator\OneDrive - Voxx Products\Desktop\Reports\Voxx_Inventory.csv"
output_dir = r"\\TORCPSER2\Public\INVENTORY"


def get_latest_snapshot_file(reports_folder_path: str, master_file_path: str) -> str:
    """
    Returns the most recently created snapshot CSV file, excluding the master file.
    """
    list_of_files = [
        file_path
        for file_path in glob.glob(os.path.join(reports_folder_path, "*.csv"))
        if os.path.abspath(file_path).lower() != os.path.abspath(master_file_path).lower()
    ]

    if not list_of_files:
        raise FileNotFoundError(f"No snapshot CSV files found in: {reports_folder_path}")

    return max(list_of_files, key=os.path.getctime)


def normalize_code(series: pd.Series) -> pd.Series:
    """
    Normalizes a Pandas Series containing string data by performing the following operations:
    - Converts all elements to string type.
    - Strips leading and trailing whitespace from each string.
    - Removes ".0" suffix from strings, if present.

    :param series: The input Pandas Series to be normalized.
    :type series: pd.Series
    :return: A normalized Pandas Series with the transformations applied.
    :rtype: pd.Series
    """
    return (
        series.astype("string")
        .str.strip()
        .str.replace(r"\.0$", "", regex=True)
    )



def read_snapshot_csv(file_path: str, clean_columns: list[str]) -> pd.DataFrame:
    """
    Reads a snapshot CSV file and returns its contents as a Pandas DataFrame. The function ensures
    the file exists, is a CSV file, and processes only specified columns.

    :param file_path: The path to the snapshot CSV file to be read.
    :type file_path: str
    :param clean_columns: A list of clean column names to be applied during reading the file.
    :type clean_columns: list[str]
    :return: A Pandas DataFrame containing the relevant data from the snapshot CSV file.
    :rtype: pandas.DataFrame
    :raises FileNotFoundError: If the specified file does not exist.
    :raises ValueError: If the file is not a CSV file.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Snapshot file not found: {file_path}")

    _, ext = os.path.splitext(file_path)
    if ext.lower() != ".csv":
        raise ValueError(f"Snapshot file must be a .csv file, got: '{ext}'")

    return pd.read_csv(
        file_path,
        skiprows=8,
        usecols=[0, 1, 2, 3, 4, 5],
        names=clean_columns
    )



def read_master_file(file_path: str) -> pd.DataFrame:
    """
    Reads a master data file and returns its content as a DataFrame.

    This function ensures the provided file path exists and checks whether the file has
    a valid Excel extension (.xlsx or .xlsm). It reads the file using the 'openpyxl' engine
    and returns the content as a Pandas DataFrame.

    :param file_path: The path to the master data file.
    :type file_path: str
    :return: A Pandas DataFrame containing the content of the master data file.
    :rtype: pd.DataFrame
    :raises FileNotFoundError: If the specified file does not exist.
    :raises ValueError: If the file extension is not .csv.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Master data file not found: {file_path}")

    _, ext = os.path.splitext(file_path)
    if ext.lower() != ".csv":
        raise ValueError(f"Master file must be a .csv file, got: '{ext}'")

    return pd.read_csv(file_path)



def format_output_excel(file_path: str) -> None:
    """
    Formats and styles an Excel file by applying font styles, fills, and borders to specified columns and
    headers. This function modifies the Excel file in place, enhancing the appearance and readability
    of the data based on predefined rules.

    :param file_path: The path to the Excel file to be formatted.
    :type file_path: str
    :return: None
    """
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    worksheet.title = "Voxx_Inventory"

    blue_fill = PatternFill(fill_type="solid", start_color="C0E6F5", end_color="C0E6F5")
    yellow_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
    bold_font = Font(bold=True)
    regular_font = Font(bold=False)
    border_color = "000000"
    border_style = "thin"
    border = Border(left=Side(style=border_style, color=border_color),
                    right=Side(style=border_style, color=border_color),
                    top=Side(style=border_style, color=border_color),
                    bottom=Side(style=border_style, color=border_color))

    center_align = Alignment(horizontal="center")

    headers = {}
    for cell in worksheet[1]:
        headers[cell.value] = cell.column
        cell.fill = blue_fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = center_align

    item_col = headers.get("Item")
    article_col = headers.get("Article")
    total_col = headers.get("TOTAL")

    for row in range(2, worksheet.max_row + 1):
        if item_col is not None:
            worksheet.cell(row=row, column=item_col).font = bold_font
            worksheet.cell(row=row, column=item_col).alignment = center_align

        if article_col is not None:
            worksheet.cell(row=row, column=article_col).font = bold_font
            worksheet.cell(row=row, column=article_col).alignment = center_align
            worksheet.cell(row=row, column=article_col).number_format = "0"

        if total_col is not None:
            total_cell = worksheet.cell(row=row, column=total_col)
            total_cell.fill = yellow_fill
            total_cell.font = regular_font
            total_cell.alignment = center_align

    for col in range(1, worksheet.max_column + 1):
        for row in range(1, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col).border = border
            worksheet.cell(row=row, column=col).alignment = center_align

    # Left align col 1
    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    worksheet.column_dimensions["A"].width = 23

    workbook.save(file_path)


def main() -> None:
    if not os.path.isdir(output_dir):
        raise FileNotFoundError(
            f"Output directory is not reachable from this run context: {output_dir}"
        )

    daily_snapshot = get_latest_snapshot_file(reports_folder, master_data_path)
    print(f"Using latest snapshot file: {daily_snapshot}")
    print("Loading data...")

    today_str = datetime.today().strftime("%#m-%#d-%Y")
    output_filename = f"{today_str} Inventory_NEW.xlsx"
    output_filepath = os.path.join(output_dir, output_filename)

    clean_columns = ["Item_Desc", "DT Code", "CA_Qty", "TN_Qty", "TX_Qty", "Total_Qty"]

    df_snapshot = read_snapshot_csv(daily_snapshot, clean_columns)
    df_snapshot = df_snapshot.dropna(subset=["DT Code"]).copy()

    df_master = read_master_file(master_data_path)

    required_snapshot_cols = {"DT Code", "Item_Desc", "CA_Qty", "TN_Qty", "TX_Qty", "Total_Qty"}
    required_master_cols = {"Article", "Item", "Brand", "MSRP"}

    missing_snapshot = required_snapshot_cols - set(df_snapshot.columns)
    missing_master = required_master_cols - set(df_master.columns)

    if missing_snapshot:
        raise ValueError(f"Missing required columns in snapshot file: {sorted(missing_snapshot)}")

    if missing_master:
        raise ValueError(f"Missing required columns in master file: {sorted(missing_master)}")

    df_snapshot["DT Code"] = normalize_code(df_snapshot["DT Code"])
    df_master["Article"] = normalize_code(df_master["Article"])

    df_snapshot = df_snapshot[df_snapshot["DT Code"].notna() & (df_snapshot["DT Code"] != "")]
    df_master = df_master[df_master["Article"].notna() & (df_master["Article"] != "")]

    print("Matching data...")
    merged_df = pd.merge(
        df_master,
        df_snapshot,
        how="left",
        left_on="Article",
        right_on="DT Code"
    )

    final_df = merged_df[["Item", "Article", "Brand", "MSRP", "Total_Qty", "CA_Qty", "TX_Qty", "TN_Qty"]].copy()
    final_df.columns = ["Item", "Article", "Brand", "MSRP", "TOTAL", "CA", "TX", "TN"]

    text_columns = ["Item", "Article", "Brand"]
    for col in text_columns:
        final_df[col] = final_df[col].fillna("")

    numeric_columns = ["MSRP", "TOTAL", "CA", "TX", "TN"]
    for col in numeric_columns:
        final_df[col] = pd.to_numeric(final_df[col], errors="coerce").fillna(0)

    final_df["Article"] = final_df["Article"].astype("Int64")

    print(f"Saving data to {output_filepath}...")
    final_df.to_excel(output_filepath, index=False)

    csv_filename = "Voxx_Inventory.csv"
    csv_filepath = os.path.join(output_dir, csv_filename)

    print(f"Overwriting CSV file at {csv_filepath}...")
    final_df.to_csv(csv_filepath, index=False)

    print("Applying Excel formatting...")
    format_output_excel(output_filepath)

    print("Copying final file to OneDrive...")

    # Creates a subfolder in OneDrive Reports folder so it doesn't get mixed up with the raw NetSuite CSVs
    onedrive_export_dir = r"C:\Users\Administrator\OneDrive - Voxx Products\Desktop\Reports\Completed_Inventory"

    os.makedirs(onedrive_export_dir, exist_ok=True)  # Forces the folder to exist

    onedrive_filepath = os.path.join(onedrive_export_dir, output_filename)
    shutil.copy2(output_filepath, onedrive_filepath)

    print("Automation complete! :P")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}")
        traceback.print_exc()
        raise